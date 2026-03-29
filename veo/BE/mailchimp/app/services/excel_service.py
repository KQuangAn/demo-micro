import asyncio
import logging
import os
from datetime import datetime, timezone
from pathlib import Path

import openpyxl
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

from app.models.webhook import ExcelRow, MailchimpWebhookPayload

logger = logging.getLogger(__name__)

HEADERS = [
    "Timestamp (UTC)",
    "Event Type",
    "Email",
    "List ID",
    "Campaign ID",
    "Campaign Subject",
    "URL (click)",
    "IP Address",
    "User Agent",
    "Reason",
]

HEADER_FILL = PatternFill(start_color="1F4E79", end_color="1F4E79", fill_type="solid")
HEADER_FONT = Font(color="FFFFFF", bold=True, size=11)

# asyncio lock — one write at a time, no file corruption
_excel_lock = asyncio.Lock()


def _init_workbook(path: Path) -> None:
    """Create a new Excel workbook with styled headers if it doesn't exist."""
    path.parent.mkdir(parents=True, exist_ok=True)

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Mailchimp Events"

    # Write and style headers
    for col_idx, header in enumerate(HEADERS, start=1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = Alignment(horizontal="center", vertical="center")

    # Set sensible column widths
    column_widths = [22, 14, 32, 18, 18, 36, 42, 18, 52, 20]
    for col_idx, width in enumerate(column_widths, start=1):
        ws.column_dimensions[get_column_letter(col_idx)].width = width

    # Freeze the header row
    ws.freeze_panes = "A2"

    wb.save(path)
    logger.info("Initialized new Excel workbook at %s", path)


def _payload_to_row(payload: MailchimpWebhookPayload) -> ExcelRow:
    """Map a webhook payload to a flat ExcelRow."""
    return ExcelRow(
        timestamp=datetime.now(timezone.utc).strftime("%Y-%m-%d %H:%M:%S"),
        event_type=payload.type.value,
        email=payload.email or "",
        list_id=payload.list_id or "",
        campaign_id=payload.campaign_id or "",
        campaign_subject=payload.campaign_subject or "",
        url=payload.url or "",
        ip=payload.ip or "",
        user_agent=payload.user_agent or "",
        reason=payload.reason or "",
    )


def _append_row_sync(file_path: str, row: ExcelRow) -> None:
    """Synchronous Excel append — runs inside asyncio executor to avoid blocking."""
    path = Path(file_path)

    if not path.exists():
        _init_workbook(path)

    wb = openpyxl.load_workbook(path)
    ws = wb.active

    ws.append([
        row.timestamp,
        row.event_type,
        row.email,
        row.list_id,
        row.campaign_id,
        row.campaign_subject,
        row.url,
        row.ip,
        row.user_agent,
        row.reason,
    ])

    wb.save(path)
    logger.info(
        "Appended %s event for %s to %s (row %d)",
        row.event_type,
        row.email or "unknown",
        path,
        ws.max_row,
    )


async def append_event_to_excel(
    payload: MailchimpWebhookPayload,
    file_path: str,
) -> None:
    """
    Thread-safe async append of a webhook event to the Excel file.
    Uses asyncio lock to prevent concurrent writes and runs
    openpyxl in a thread executor to avoid blocking the event loop.
    """
    row = _payload_to_row(payload)

    async with _excel_lock:
        loop = asyncio.get_running_loop()
        await loop.run_in_executor(None, _append_row_sync, file_path, row)
